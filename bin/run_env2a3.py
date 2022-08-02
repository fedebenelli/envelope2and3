#!/usr/bin/env python3
"""
"""
from glob import glob
import os
import shutil
from openpyxl import load_workbook
import sys
import matplotlib.pyplot as plt
from math import log
import platform


def get_concentration():
    """ From an envelIN file read the concentrations and return them as a list
    """
    with open('envelIN.txt', encoding='utf-8') as f:
        concentrations = f.readlines()[1].replace('\t', ' ')

    concentrations = [float(zi) for zi in concentrations.split(' ')]
    concentrations = [round(zi, 4) for zi in concentrations]
    concentrations = [str(zi) for zi in concentrations]

    return concentrations


def iter_sheet(ws):
    """Iterate in an Excel WorkSheet and extract data"""
    on_table = False
    for i, row in enumerate(ws.iter_rows()):
        row3_val = row[3].value
        print(row3_val)

        try:
            float(row3_val)
        except:
            if row3_val is None:
                on_table = False
                continue
            if not on_table and row3_val.startswith("Modi"):
                on_table = True
                n = int(row[0].value)
                header = row3_val
                os.mkdir(os.path.join(INFILES_PATH, header))
            continue

        if on_table:
            label = row[0].value
            if "Cambio %" in label or "Origin" in label:
                continue

            values = []
            for cell in row[1:n+1]:
                value = cell.value
                values.append(str(round(value, 10)))

                # first_decimal = -log(value)/log(10)
                # if first_decimal > 4:
                #    values.append(str(round(value, first_decimal+2)))
                # else:
                #    values.append(str(round(value, 4)))
            
            # Read original file
            with open(BASE_CASE, "r") as f:
                original_lines = f.readlines()
                original_lines[1] = ' '.join(values)
                original_lines[1] += '\n'

            # Write new file
            with open(
                    os.path.join(INFILES_PATH, header, f"{label}"), "w"
                    ) as w:
                w.write(''.join(original_lines))


def gen_infiles(file, worksheet=None):
    wb = load_workbook(file, data_only=True)
    if worksheet:
        sheets = [wb[worksheet]]
    else:
        sheets = wb.worksheets
    for ws in sheets:
        iter_sheet(ws)


def run_fortran(i, out_dir):
    run_str = ''
    if EXEC.endswith('exe') and OS == "Linux":
        run_str += f"wine '{EXEC}'"
    else:
        run_str = EXEC

    os.system(f'"{run_str}"')

    # make_envel()
    envelopes = sorted(glob("envelout*"))

    make_figure(envelopes, title='\n'.join(out_dir.split("/")[-2:]))

    if i != 0:
        shutil.copy(
                'envelopes.png',
                os.path.join(OUTFILES_PATH, f"envelopes/{str(i).zfill(3)}.png")
                )
    shutil.move('envelopes.png', out_dir)
    for envel in envelopes:
        shutil.move(envel, out_dir)


def plot_envelope(file, label, ax):
    ts = []
    ps = []

    if label == 'dew':
        color = 'blue'
    elif label == 'bub':
        color = 'black'
    else:
        color = None

    with open(file) as f:
        for line in f.readlines()[1:]:
            if line.split() == []:
                break
            x, y = line.split()[:2]
            ts.append(float(x))
            ps.append(float(y))

    ax.plot(ts, ps, label=label, color=color)


def show_crit(file, label, ax):
    """ Extract the critical points from an envelOUT file and add them to a
    plot."""
    with open(file) as f:
        lines = f.readlines()

        n_crit = 0
        i = 0

        for i, line in enumerate(lines):
            try:
                if line.split()[0].startswith('Number'):
                    n_crit = int(line.split()[-1])
                    break
            except IndexError:
                pass

        for line in lines[i+2:i+2+n_crit]:
            x, y = line.split()[:2]
            x, y = float(x), float(y)
            ax.scatter(x, y, color='black')


def make_figure(envelopes, title):
    """ From a set of envelopes plot them. """
    plt.clf()
    for i, file in enumerate(envelopes):
        ax = plt.subplot()
        show_crit(file, '', ax)
        if i == 0:
            plot_envelope(file, 'dew', ax)
        elif i == 1:
            plot_envelope(file, 'bub', ax)
        elif i > 1:
            plot_envelope(file, i-1, ax)

    concentrations = get_concentration()
    plt.annotate(
            '\n'.join(concentrations),
            xy=(1.01, 0.2), xycoords='axes fraction')
    plt.title(title)
    plt.legend()
    plt.xlim(200, 900)
    plt.ylim(0, 350)
    plt.savefig("envelopes.png", dpi=200)
    plt.show()

    return plt.gcf()


def single(file="envelIN.txt"):
    """Run a single envelIN file"""


def clean():
    """Clean generated files"""
    os.remove(glob("envelo*"))


def excel():
    """Run the cases included in an Excel file.
    """




if __name__ == '__main__':
    EXEC = os.path.abspath(sys.argv[1])
    PATH = os.path.normpath(os.path.join(__file__, os.pardir))
    INFILES_PATH = os.path.join(PATH, 'infiles')
    OUTFILES_PATH = os.path.join(PATH, 'outfiles')
    BASE_CASE = "study_case"
    OS = platform.system()

    CASES = {
                "single": single,
                "excel": excel,
            }

    if len(sys.argv) > 2:
        if sys.argv[2] == "single":
            print("running single")
            run_fortran(0, "singleout")

        
        elif sys.argv[2] == 'full':
            # Remove all runs
            shutil.rmtree("outfiles")
            shutil.rmtree("infiles")

            # Make room for new runs
            os.mkdir('infiles')
            os.mkdir('outfiles')
            os.mkdir('outfiles/envelopes')

            # First run the original case
            shutil.copy(BASE_CASE, "envelIN.txt")
            original_outdir = os.path.join(OUTFILES_PATH, "original")
            os.mkdir(original_outdir)
            run_fortran(0, original_outdir)

            # Check if the user wants to run a specific worksheet
            try:
                file, worksheet = sys.argv[3:5]
                if worksheet == "all":
                    worksheet = None
            except IndexError:
                print("Selecionar archivo y hoja de trabajo")
                exit()

            gen_infiles(file, worksheet)

            i = 0
            # Iterate through each infile folder
            for folder in os.listdir(INFILES_PATH):
                # Iterate through each infile
                for file in sorted(
                        os.listdir(os.path.join(INFILES_PATH, folder))
                        ):
                    i += 1
                    file_abs = os.path.join(INFILES_PATH, folder, file)
                    results_dir = os.path.join(OUTFILES_PATH, folder, file)

                    try:
                        os.mkdir(os.path.join(OUTFILES_PATH, folder))
                    except FileExistsError:
                        ...
                    os.mkdir(results_dir)

                    print("Working on: ", results_dir.split("/")[-1])

                    shutil.copy(file_abs, "envelIN.txt")
                    run_fortran(i, results_dir)
